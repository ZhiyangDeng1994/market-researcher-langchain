from pathlib import Path
from openpyxl import Workbook
from openpyxl.comments import Comment


def build_comps_xlsx(rows, sector: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comps"

    ws["A1"] = f"{sector.upper()} — COMPARABLE COMPANY ANALYSIS"
    ws["A2"] = "All figures in USD millions except per-share amounts and ratios"

    headers = ["Ticker", "EV ($mm)", "EBITDA ($mm)", "Price", "EPS", "EV/EBITDA", "P/E"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)

    start = 5
    for i, r in enumerate(rows):
        row = start + i
        ws.cell(row=row, column=1, value=r.ticker)
        for col, val in [(2, r.ev), (3, r.ebitda), (4, r.price), (5, r.eps)]:
            c = ws.cell(row=row, column=col, value=val)       # raw input
            c.comment = Comment(f"Source: {r.source}", "agent")  # source on every input
        ws.cell(row=row, column=6, value=f"=B{row}/C{row}")   # formula, not hardcode
        ws.cell(row=row, column=7, value=f"=D{row}/E{row}")

    end = start + len(rows) - 1
    stats = [("Mean", "AVERAGE"), ("Median", "MEDIAN"), ("Min", "MIN"), ("Max", "MAX")]
    for j, (label, fn) in enumerate(stats):
        row = end + 2 + j
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=6, value=f"={fn}(F{start}:F{end})")
        ws.cell(row=row, column=7, value=f"={fn}(G{start}:G{end})")

    out = Path("out"); out.mkdir(exist_ok=True)
    path = out / "comps.xlsx"
    wb.save(path)
    return str(path)