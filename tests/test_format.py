"""Output quality tests — deterministic, zero-cost, no API calls.
Run:  pytest tests/test_format.py -v
Requires: python main.py has been run at least once.
"""
import json
import re
from pathlib import Path

import pytest
from openpyxl import load_workbook

# ── Anchor all paths to project root ──
ROOT = Path(__file__).resolve().parent.parent
MANIFEST = ROOT / "out" / "manifest.json"


@pytest.fixture(scope="module")
def manifest():
    assert MANIFEST.exists(), (
        "out/manifest.json not found — run `python main.py` first"
    )
    return json.loads(MANIFEST.read_text())


@pytest.fixture(scope="module")
def primer(manifest):
    path = ROOT / manifest["primer"]
    assert path.exists(), f"Primer not found: {path}"
    return path.read_text(encoding="utf-8")


@pytest.fixture(scope="module")
def xlsx(manifest):
    raw = manifest.get("comps_xlsx", "")
    if not raw:
        # fallback: glob for any xlsx in out/
        candidates = sorted((ROOT / "out").glob("comps*.xlsx"),
                            key=lambda f: f.stat().st_mtime)
        assert candidates, "No comps xlsx found in out/"
        return candidates[-1]
    path = ROOT / raw
    assert path.exists(), f"Excel not found: {path}"
    return path


# ═══════════════════════════════════════════
#  Markdown — structure
# ═══════════════════════════════════════════

def test_all_sections_present(primer):
    for section in ["## Overview", "## Comps", "## Ideas"]:
        assert section in primer, f"Missing section: {section}"


def test_has_title(primer):
    assert primer.startswith("# "), "Primer should start with a top-level heading"


# ═══════════════════════════════════════════
#  Markdown — Overview
# ═══════════════════════════════════════════

def _overview_bullets(primer: str) -> list[str]:
    overview = primer.split("## Overview")[1].split("## Comps")[0]
    return [l.strip() for l in overview.splitlines() if l.strip().startswith("- ")]


def test_overview_has_enough_facts(primer):
    bullets = _overview_bullets(primer)
    assert len(bullets) >= 8, (
        f"Overview has only {len(bullets)} facts, expected >= 8"
    )


def test_overview_facts_have_sources(primer):
    for line in _overview_bullets(primer):
        assert "—" in line and "*" in line, (
            f"Fact missing source marker: {line[:80]}..."
        )


def test_overview_facts_have_numbers(primer):
    num_pattern = re.compile(r"\d[\d,.]*\s?(%|x|GW|MW|TWh|bn|billion|million|\$)", re.I)
    bullets = _overview_bullets(primer)
    with_numbers = [b for b in bullets if num_pattern.search(b)]
    assert len(with_numbers) >= len(bullets) // 2, (
        f"Only {len(with_numbers)}/{len(bullets)} facts have numbers"
    )


# ═══════════════════════════════════════════
#  Markdown — Comps
# ═══════════════════════════════════════════

def _comps_section(primer: str) -> str:
    return primer.split("## Comps")[1].split("## Ideas")[0]


def test_comps_has_table(primer):
    assert "|" in _comps_section(primer), "Comps section has no markdown table"


def test_comps_has_multiple_tickers(primer):
    rows = [l for l in _comps_section(primer).splitlines() if l.strip().startswith("|")]
    data_rows = [r for r in rows if not re.match(r"^\|\s*-", r)]
    assert len(data_rows) >= 5, f"Comps table only {len(data_rows)} rows, expected >= 5"


def test_comps_no_uniform_stub(primer):
    section = _comps_section(primer)
    assert section.count("12.5x") <= 1, "Likely stub data (12.5x repeated)"


# ═══════════════════════════════════════════
#  Markdown — Ideas
# ═══════════════════════════════════════════

def _ideas_section(primer: str) -> str:
    return primer.split("## Ideas")[1]


def test_ideas_not_empty(primer):
    assert len(_ideas_section(primer).strip()) > 500, "Ideas section too short"


def test_ideas_have_multiple_names(primer):
    ideas = _ideas_section(primer)
    ticker_pattern = re.compile(r"\b[A-Z]{2,5}\b")
    tickers = set(ticker_pattern.findall(ideas))
    noise = {"THE", "AND", "FOR", "NOT", "BUT", "WITH", "THIS", "THAT", "FROM",
             "LONG", "SHORT", "KEY", "USD", "EPS", "EBITDA", "GW", "MW", "TWH",
             "NM", "OEM", "EPC", "PPA", "IRA", "ROE", "FCF", "CAGR", "IPP"}
    real = tickers - noise
    assert len(real) >= 3, f"Only found {len(real)} likely tickers: {real}"


def test_ideas_mention_risks(primer):
    assert "risk" in _ideas_section(primer).lower(), "Ideas should discuss risks"


# ═══════════════════════════════════════════
#  Markdown — Guardrail
# ═══════════════════════════════════════════

def test_no_unsourced_overview_facts(primer):
    for line in _overview_bullets(primer):
        assert "[UNSOURCED]" not in line, f"Overview fact failed guardrail: {line[:80]}"


# ═══════════════════════════════════════════
#  Excel — structure
# ═══════════════════════════════════════════

def test_xlsx_has_header_row(xlsx):
    ws = load_workbook(xlsx).active
    headers = [ws.cell(row=4, column=c).value for c in range(1, 8)]
    assert "Ticker" in headers, f"Expected 'Ticker' in headers, got: {headers}"


def test_xlsx_has_data_rows(xlsx):
    ws = load_workbook(xlsx).active
    tickers = []
    for row in range(5, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and val not in ("Mean", "Median", "Min", "Max"):
            tickers.append(val)
    assert len(tickers) >= 3, f"Only {len(tickers)} data rows"


# ═══════════════════════════════════════════
#  Excel — formulas
# ═══════════════════════════════════════════

def test_xlsx_ev_ebitda_is_formula(xlsx):
    cell = load_workbook(xlsx).active.cell(row=5, column=6).value
    assert isinstance(cell, str) and cell.startswith("="), (
        f"EV/EBITDA (F5) should be formula, got: {cell}"
    )


def test_xlsx_pe_is_formula(xlsx):
    cell = load_workbook(xlsx).active.cell(row=5, column=7).value
    assert isinstance(cell, str) and cell.startswith("="), (
        f"P/E (G5) should be formula, got: {cell}"
    )


def test_xlsx_summary_stats_are_formulas(xlsx):
    ws = load_workbook(xlsx).active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Mean":
            f_val = ws.cell(row=row, column=6).value
            assert isinstance(f_val, str) and "AVERAGE" in f_val.upper(), (
                f"Mean EV/EBITDA should be AVERAGE formula, got: {f_val}"
            )
            return
    pytest.fail("No 'Mean' row found")


# ═══════════════════════════════════════════
#  Excel — source comments
# ═══════════════════════════════════════════

def test_xlsx_input_cells_have_comments(xlsx):
    ws = load_workbook(xlsx).active
    missing = []
    for row in range(5, 8):
        for col in (2, 3):
            if ws.cell(row=row, column=col).comment is None:
                missing.append(ws.cell(row=row, column=col).coordinate)
    assert not missing, f"Input cells missing source comments: {missing}"


def test_xlsx_comment_mentions_source(xlsx):
    comment = load_workbook(xlsx).active.cell(row=5, column=2).comment
    assert comment is not None, "First EV cell has no comment"
    text = comment.text.lower()
    assert "source" in text or "yahoo" in text or "yfinance" in text, (
        f"Comment doesn't mention source: {comment.text}"
    )